import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
import io
import datetime
import requests
import urllib3
import base64
import zipfile  
import re       
import time     
from streamlit_cropper import st_cropper
import openpyxl # 🌟 엑셀 원본 서식 제어용 라이브러리
from openpyxl.styles import Alignment, Border, Side, Font

# --- [네트워크 보안 경고 무시 설정] ---
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- [초기 설정] ---
st.set_page_config(page_title="중구의회 언론보도 스크랩", layout="wide")

# =====================================================================
# 🔒 [보안 설정] 팀원들과 공유할 웹사이트 접속 비밀번호
APP_PASSWORD = "2904220" 
# =====================================================================

# --- [로그인 화면 로직] ---
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.title("🔒 중구의회 언론보도 스크랩 시스템")
    st.info("울산 중구의회 정책이들의 전용 시스템입니다. 비밀번호를 입력해 주세요.")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pwd_input = st.text_input("비밀번호", type="password")
        if st.button("로그인", use_container_width=True):
            if pwd_input == APP_PASSWORD:
                st.session_state.logged_in = True
                st.rerun() 
            else:
                st.error("❌ 비밀번호가 일치하지 않습니다.")
    st.stop() 

# --- [상태 초기화 및 사전 데이터] ---
if 'scraped_data' not in st.session_state: st.session_state.scraped_data = []
if 'newly_added_data' not in st.session_state: st.session_state.newly_added_data = [] 
if 'original_excel_bytes' not in st.session_state: st.session_state.original_excel_bytes = None 
if 'current_page' not in st.session_state: st.session_state.current_page = 0
if 'excel_loaded' not in st.session_state: st.session_state.excel_loaded = False

categories = ["경제.민생", "기타", "문화", "행사", "보건", "복지", "안전.교육", "의원연구단체", "쟁점사항", "주요사업", "환경", "관광"]
committees = ["행정자치위원회", "복지건설위원회", "의회운영위원회", "행정자치.복지건설위원회", "기타"]
press_list = ["경상일보", "동아일보", "서울경제", "연합뉴스", "울산광역매일", "울산매일", "울산신문", "울산제일일보", "조선일보", "중앙일보", "한국경제"]

dept_to_committee = {
    "기획예산실": "행정자치위원회", "홍보실": "행정자치위원회", "경제정책과": "행정자치위원회", 
    "전통시장과": "행정자치위원회", "문화관광과": "행정자치위원회", "일자리정책과": "행정자치위원회", 
    "총무과": "행정자치위원회", "자치행정과": "행정자치위원회", "회계과": "행정자치위원회", 
    "세무1과": "행정자치위원회", "세무2과": "행정자치위원회", "보건소": "행정자치위원회", 
    "보건과": "행정자치위원회", "건강관리과": "행정자치위원회", "문화의전당": "행정자치위원회", 
    "구립도서관": "행정자치위원회", "도시관리공단": "행정자치위원회",
    "복지지원과": "복지건설위원회", "노인장애인과": "복지건설위원회", "가족복지과": "복지건설위원회", 
    "교육체육과": "복지건설위원회", "교통과": "복지건설위원회", "공원녹지과": "복지건설위원회", 
    "환경위생과": "복지건설위원회", "환경미화과": "복지건설위원회", "공간정보과": "복지건설위원회", 
    "안전총괄과": "복지건설위원회", "건설과": "복지건설위원회", "도시과": "복지건설위원회", 
    "건축과": "복지건설위원회", "시설지원과": "복지건설위원회",
    "의회사무국": "의회운영위원회"
}

def get_committee_by_dept(dept_name):
    for key, value in dept_to_committee.items():
        if key in dept_name: return value
    return "기타"

# =====================================================================
# 🚨 [수정 완료] 스트림릿 금고(Secrets)에서 API 키를 안전하게 꺼내옵니다! 
API_KEY = st.secrets["API_KEY"]
# =====================================================================

def get_best_available_model(api_key):
    url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
    headers = {'x-goog-api-key': api_key} 
    try:
        res = requests.get(url, headers=headers, verify=False)
        if res.status_code == 200:
            models = res.json().get('models', [])
            valid_models = [m['name'] for m in models if 'generateContent' in m.get('supportedGenerationMethods', [])]
            if not valid_models: return None, "사용 가능한 모델이 없습니다."
            for target in ["models/gemini-1.5-flash", "models/gemini-1.5-pro", "models/gemini-1.0-pro-vision"]:
                if target in valid_models: return target, "성공"
            return valid_models[0], "성공"
        else: return None, f"에러: {res.status_code}"
    except Exception as e: return None, f"에러: {str(e)}"

# --- [사이드바] ---
st.sidebar.header("⚙️ 날짜 설정 및 파일 업로드")
report_date = st.sidebar.text_input("보도일자 (yymmdd)", value=datetime.datetime.now().strftime("%y%m%d"))

st.sidebar.markdown("---")
uploaded_excel = st.sidebar.file_uploader("📁 어제까지 정리된 엑셀 업로드 (선택)", type=['xlsx', 'xls'])

if uploaded_excel is not None and not st.session_state.excel_loaded:
    try:
        st.session_state.original_excel_bytes = uploaded_excel.getvalue() 
        old_df = pd.read_excel(uploaded_excel)
        st.session_state.scraped_data = old_df.to_dict('records')
        st.session_state.excel_loaded = True
        st.sidebar.success(f"✅ 기존 엑셀 원본 서식 보존 및 {len(st.session_state.scraped_data)}건 불러오기 완료!")
    except Exception as e:
        st.sidebar.error(f"엑셀 파일 읽기 오류: {e}")

st.sidebar.markdown("---")
uploaded_file = st.sidebar.file_uploader("📄 오늘 주요언론보도 PDF 업로드", type=['pdf'])

# --- [메인 로직] ---
st.title("📰 일일 언론보도 스크랩 자동화 시스템")

if 'ai_title' not in st.session_state: st.session_state.ai_title = ""
if 'ai_press_idx' not in st.session_state: st.session_state.ai_press_idx = 0
if 'ai_category_idx' not in st.session_state: st.session_state.ai_category_idx = 0
if 'ai_committee_idx' not in st.session_state: st.session_state.ai_committee_idx = 0
if 'ai_dept' not in st.session_state: st.session_state.ai_dept = ""
if 'ai_summary' not in st.session_state: st.session_state.ai_summary = ""

if uploaded_file:
    doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
    total_pages = len(doc)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if st.button("⬅️ 이전 페이지") and st.session_state.current_page > 0:
            st.session_state.current_page -= 1
    with col2:
        st.write(f"**현재 페이지: {st.session_state.current_page + 1} / {total_pages}**")
    with col3:
        if st.button("다음 페이지 ➡️") and st.session_state.current_page < total_pages - 1:
            st.session_state.current_page += 1

    page = doc.load_page(st.session_state.current_page)
    pix = page.get_pixmap(dpi=200)
    img_data = pix.tobytes("png")
    image = Image.open(io.BytesIO(img_data))

    left_col, right_col = st.columns([1.2, 1])

    with left_col:
        st.info("✂️ **마우스로 기사 영역을 드래그하여 지정하세요.**")
        cropped_image = st_cropper(image, realtime_update=True, box_color='blue', aspect_ratio=None)

    with right_col:
        st.subheader("📝 기사 정보 입력")
        st.image(cropped_image, width=300, caption="선택된 기사 미리보기")
        
        if st.button("✨ Gemini로 기사 분석하기"):
            with st.spinner("Gemini가 이미지를 분석 중입니다. 서버 혼잡 시 자동으로 대기 후 재시도합니다..."):
                best_model, status_msg = get_best_available_model(API_KEY)
                
                if best_model:
                    try:
                        buffered = io.BytesIO()
                        cropped_image.save(buffered, format="PNG")
                        img_str = base64.b64encode(buffered.getvalue()).decode("utf-8")

                        all_dept_list_str = ", ".join(dept_to_committee.keys())
                        all_category_str = ", ".join(categories)

                        prompt = f"""
                        다음 이미지는 울산광역시 중구 관련 언론보도 스크랩입니다. 이미지 내용을 분석해서 아래 정보를 정확히 추출해주세요.
                        1. 기사 제목 (Title)
                        2. 언론사: 이미지 좌상단 등에서 신문사 이름을 찾아 다음 중 하나만 출력하세요. ({', '.join(press_list)})
                        3. 예상 담당부서: 반드시 다음 목록 중에서 가장 관련 있는 부서 '1개'의 이름만 정확히 출력하세요. 없으면 빈칸으로 두세요. (목록: {all_dept_list_str})
                        4. 구분: 기사가 다루는 내용을 파악하여 반드시 다음 목록 중 가장 알맞은 '1개'만 정확히 출력하세요. (목록: {all_category_str})
                        5. 주요내용: 기사의 핵심을 중구청 위주로 '반드시 개조식(- 기호 시작)'으로 한 문장씩 3~4줄로 요약하세요. 
                           ▶ [중요] 각 문장의 끝은 반드시 '~함', '~할 계획임', '~할 예정임', '~개최함' 과 같이 간결한 명사형 종결어미로 끝내세요.

                        출력형식:
                        제목: [제목]
                        언론사: [언론사명]
                        부서: [부서명]
                        구분: [카테고리명]
                        요약:
                        - [요약문 1]
                        - [요약문 2]
                        """
                        
                        url = f"https://generativelanguage.googleapis.com/v1beta/{best_model}:generateContent?key={API_KEY}"
                        headers = {'Content-Type': 'application/json', 'x-goog-api-key': API_KEY}
                        data = {
                            "contents": [{"parts": [{"text": prompt}, {"inlineData": {"mimeType": "image/png", "data": img_str}}]}]
                        }
                        
                        max_retries = 3
                        retry_delay = 5
                        
                        for attempt in range(max_retries):
                            response = requests.post(url, headers=headers, json=data, verify=False)
                            
                            if response.status_code == 200:
                                result = response.json()
                                response_text = result['candidates'][0]['content']['parts'][0]['text']
                                
                                lines = response_text.split('\n')
                                parsed_title, parsed_press, parsed_dept, parsed_category = "", "", "", ""
                                is_summary = False
                                summary_lines = []
                                
                                for line in lines:
                                    line = line.strip()
                                    if not line: continue
                                    if line.startswith("제목:"): parsed_title = line.replace("제목:", "").strip()
                                    elif line.startswith("언론사:"): parsed_press = line.replace("언론사:", "").strip()
                                    elif line.startswith("부서:"): parsed_dept = line.replace("부서:", "").strip()
                                    elif line.startswith("구분:"): parsed_category = line.replace("구분:", "").strip()
                                    elif line.startswith("요약:"): is_summary = True
                                    elif is_summary: summary_lines.append(line)
                                
                                st.session_state.ai_title = parsed_title
                                st.session_state.ai_dept = parsed_dept
                                st.session_state.ai_summary = "\n".join(summary_lines)
                                
                                if parsed_press in press_list: st.session_state.ai_press_idx = press_list.index(parsed_press)
                                if parsed_category in categories: st.session_state.ai_category_idx = categories.index(parsed_category)
                                matched_committee = get_committee_by_dept(parsed_dept)
                                if matched_committee in committees: st.session_state.ai_committee_idx = committees.index(matched_committee)
                                    
                                st.toast(f"✅ Gemini 분석 성공!")
                                st.rerun()
                                break
                                
                            elif response.status_code == 429:
                                if attempt < max_retries - 1:
                                    st.warning(f"서버가 혼잡합니다. {retry_delay}초 후 자동으로 재시도합니다... (시도 횟수: {attempt+1}/{max_retries})")
                                    time.sleep(retry_delay)
                                    retry_delay += 5
                                else:
                                    st.error("❌ 일일 무료 사용량을 초과했거나 기관망 제한에 걸린 것 같습니다. 잠시 후 다시 시도해 주세요.")
                            else:
                                st.error(f"❌ 분석 실패: {response.status_code}")
                                break
                                
                    except Exception as e: st.error(f"AI 호출 오류: {e}")
                else:
                    st.error("❌ 구글 API 연결 실패: 모델을 찾을 수 없습니다.")

        with st.form("article_input_form"):
            title = st.text_input("제목", value=st.session_state.ai_title)
            press = st.selectbox("언론사", press_list, index=st.session_state.ai_press_idx)
            category = st.selectbox("구분", categories, index=st.session_state.ai_category_idx)
            committee = st.selectbox("위원회명", committees, index=st.session_state.ai_committee_idx)
            dept = st.text_input("담당부서", value=st.session_state.ai_dept)
            summary = st.text_area("주요내용", value=st.session_state.ai_summary, height=150)
            note = st.text_input("비고")
            
            submitted = st.form_submit_button("💾 스크랩 목록에 저장")
            
            if submitted:
                today_count = sum(1 for item in st.session_state.scraped_data if str(item.get('언론보도일자', '')) == str(report_date))
                current_count = today_count + 1
                serial_num = f"{report_date}-{current_count}"
                
                new_data = {
                    "연번": serial_num,
                    "구분": category,
                    "위원회명": committee,
                    "언론사": press,
                    "언론보도일자": report_date,
                    "담당부서": dept,
                    "제목": title,
                    "주요내용": summary,
                    "비고": note,
                    "이미지": cropped_image 
                }
                st.session_state.scraped_data.append(new_data)
                st.session_state.newly_added_data.append(new_data) 
                
                st.session_state.ai_title = ""
                st.session_state.ai_dept = ""
                st.session_state.ai_summary = ""
                st.session_state.ai_press_idx = 0
                st.session_state.ai_category_idx = 0
                st.session_state.ai_committee_idx = 0
                
                st.success(f"[{serial_num}] 저장 완료! (총 {len(st.session_state.scraped_data)}건 누적)")
                st.rerun()

# --- [하단: 누적 데이터 확인 및 엑셀 다운로드] ---
st.markdown("---")
st.subheader("📊 스크랩 누적 목록")

if st.session_state.scraped_data:
    df = pd.DataFrame(st.session_state.scraped_data)
    if '이미지' in df.columns:
        df_display = df.drop(columns=['이미지'])
    else:
        df_display = df
        
    st.dataframe(df_display, use_container_width=True)
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        
        excel_output = io.BytesIO()
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if st.session_state.original_excel_bytes:
            wb = openpyxl.load_workbook(io.BytesIO(st.session_state.original_excel_bytes))
            ws = wb.active
            
            for data in st.session_state.newly_added_data:
                row_val = [
                    data.get("연번", ""), data.get("구분", ""), data.get("위원회명", ""),
                    data.get("언론사", ""), data.get("언론보도일자", ""), data.get("담당부서", ""),
                    data.get("제목", ""), data.get("주요내용", ""), data.get("비고", "")
                ]
                ws.append(row_val)
                
                for cell in ws[ws.max_row]:
                    cell.border = thin_border
                    if cell.column in [7, 8]:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center

            wb.save(excel_output)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '스크랩목록'
            
            headers = ["연번", "구분", "위원회명", "언론사", "언론보도 일자", "담당부서", "제목", "주요내용", "비고"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
                
            col_widths = {'A': 12, 'B': 12, 'C': 18, 'D': 12, 'E': 12, 'F': 15, 'G': 40, 'H': 60, 'I': 10}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
                
            for data in st.session_state.scraped_data: 
                row_val = [
                    data.get("연번", ""), data.get("구분", ""), data.get("위원회명", ""),
                    data.get("언론사", ""), data.get("언론보도일자", ""), data.get("담당부서", ""),
                    data.get("제목", ""), data.get("주요내용", ""), data.get("비고", "")
                ]
                ws.append(row_val)
                for cell in ws[ws.max_row]:
                    cell.border = thin_border
                    if cell.column in [7, 8]:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                        
            wb.save(excel_output)
        
        excel_filename = f"2026년 일일 언론보도 스크랩 목록({report_date[2:6]}).xlsx"
        zf.writestr(excel_filename, excel_output.getvalue())
        
        for data in st.session_state.newly_added_data:
            if '이미지' in data and data['이미지'] is not None:
                img = data['이미지']
                safe_title = re.sub(r'[\\/*?:"<>|]', "", str(data.get('제목', ''))).strip()
                img_filename = f"({data.get('담당부서', '')})({data.get('연번', '')}){safe_title}.jpg"
                
                img_byte_arr = io.BytesIO()
                if img.mode in ("RGBA", "P"): 
                    img = img.convert("RGB")
                
                img.save(img_byte_arr, format='JPEG', quality=95)
                zf.writestr(img_filename, img_byte_arr.getvalue())
            
    zip_filename = ""
    if len(report_date) == 6:
        mm = report_date[2:4]
        dd = report_date[4:6]
        zip_filename = f"{mm}월{dd}일 주요언론보도.zip"
    else:
        zip_filename = f"{report_date}_주요언론보도.zip" 

    st.download_button(
        label="📦 엑셀 및 기사 이미지(ZIP) 일괄 다운로드",
        data=zip_buffer.getvalue(),
        file_name=zip_filename,
        mime="application/zip"
    )
    
    if st.button("🗑️ 목록 완전히 초기화하기 (로그아웃)"):
        st.session_state.scraped_data = []
        st.session_state.newly_added_data = []
        st.session_state.original_excel_bytes = None
        st.session_state.excel_loaded = False
        st.session_state.logged_in = False 
        st.rerun()
else:
    st.info("아직 저장되거나 불러온 기사가 없습니다.")